from openpyxl import load_workbook

def get_xlsx_rows_as_dicts(xlsx_file):
    rows_as_dicts = []

    sheet = load_workbook(xlsx_file).active
    columns_names = [c.value.strip() if type(c.value) == str else c.value for c in sheet[1]]
    for row in sheet.iter_rows(min_row=2):
        row_values = [c.value.strip() if type(c.value) == str else c.value for c in row]
        rows_as_dicts.append(dict(zip(columns_names, row_values)))

    return rows_as_dicts

def get_xlsx_sheet_rows_as_dicts(sheet):
    rows_as_dicts = []

    columns_names = [c.value for c in sheet[1]]
    for row in sheet.iter_rows(min_row=2):
        if row == None:
            continue
        row_values = [c.value for c in row]
        rows_as_dicts.append(dict(zip(columns_names, row_values)))

    return rows_as_dicts

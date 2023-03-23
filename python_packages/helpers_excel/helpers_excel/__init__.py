from openpyxl import load_workbook


def get_xlsx_rows_as_dicts(xlsx_file):
    data = {}

    wb = load_workbook(xlsx_file)

    for sheet in wb.worksheets:
        rows = []

        columns_names = [c.value.strip() if type(c.value) == str else c.value for c in sheet[1]]

        for row in sheet.iter_rows(min_row=2):
            row_values = [c.value.strip() if type(c.value) == str else c.value for c in row]
            rows.append(dict(zip(columns_names, row_values)))

        data[sheet.title] = rows

    return data
